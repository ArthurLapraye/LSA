#!/usr/bin/python
#-*- encoding: utf-8 -*-
#Arthur Lapraye - copyright 2016
# This file is part of X-TAL.

# X-TAL is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# X-TAL is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with X-TAL.  If not, see <http://www.gnu.org/licenses/>. 

import openpyxl as opx
import gensim
import sys
import os
import csv
import logging
import functools
import traceback
import re

from collections import defaultdict
from gensim import corpora, models, similarities,matutils #Licence LGPL (check version)


from PyQt4 import *
from PyQt4 import QtCore, QtGui, QtWebKit
from PyQt4.QtGui import *
from PyQt4.QtCore import *

_fromUtf8 = QtCore.QString.fromUtf8

from lemmtok import Lemmtok

LEFFFPATH=""

class Table(QtGui.QTableWidget):
	def __init__(self,sheet=None,dimensions=None):
		
		if sheet:
			sheet=sheet
				
			self.row_count = sheet.max_row + 1
			self.column_count = sheet.max_column + 1
			super(Table,self).__init__(self.row_count,self.column_count)
			
			
			r,y=0,0
			for row in sheet:
				y=0
				for fn in row:
					if fn.value:
						newitem=QtGui.QTableWidgetItem( _fromUtf8(unicode(fn.value)))
					else:
						newitem=QtGui.QTableWidgetItem("")
					
					self.setItem(r,y,newitem)
					
					y += 1
					
				r += 1
		elif dimensions:
			self.row_count,self.column_count=dimensions
			super(Table,self).__init__(self.row_count,self.column_count)
		else:
			super(Table,self).__init__()
		
	def __getitem__(self,pair):
		row,col=pair
		item=self.item(row, col)
		#print item
		z = unicode(item.text()) if item else ""
		
		return z
	
	def __setitem__(self,pair,value):
		row,col=pair
		
		item=self.item(row, col)
		
		if not item:
			if value:
				item=QtGui.QTableWidgetItem(value)
			else:
				item=QtGui.QTableWidgetItem("")
			self.setItem(row,col, item)
		else:
			item.setText(value)
	
	def iteritems(self):
		"""Générateur qui renvoie tout les items de haut en bas et de gauche à droite"""
		for row in xrange(0,self.row_count):
			for col in xrange(0,self.column_count):
				yield self.item(row,col)
	
	
		#QMessageBox.about(self, "Info",value)
		
class Main(QtGui.QMainWindow):
	
	def __init__(self):
		"""
		Fonction d'initialisation de la fenêtre, met en place tout les éléments interactifs du programme. 
		"""
		self.new=0
		super(Main, self).__init__()
		self.sys_clip = QtGui.QApplication.clipboard()
		self.setContextMenuPolicy(Qt.ActionsContextMenu)
		self.setWindowTitle(_fromUtf8("X-TAL"))
		self.setGeometry(100,100,800,600)
		self.show()
		
		#Actions du menu fichier
		
		exitAction = QtGui.QAction(QtGui.QIcon('exit.png'), '&Quitter', self)  
		exitAction.setShortcut('Ctrl+Q')
		exitAction.setStatusTip("Quitte l'application")
		exitAction.triggered.connect(QtGui.qApp.quit)
		
		loadAction = QtGui.QAction(QtGui.QIcon('open.png'),'&Ouvrir',self)
		loadAction.setShortcut("Ctrl+O")
		loadAction.setStatusTip("Ouvre un fichier .xlsx ou .csv")
		loadAction.triggered.connect(lambda : self.openfile(self.getfilename(flag=1) ) )
		
		
		saveAction = QtGui.QAction(QtGui.QIcon('save.png'),'&Sauvegarder',self)
		saveAction.setShortcut('Ctrl+S')
		saveAction.setStatusTip(_fromUtf8(u"Sauvegarde le fichier courant"))
		saveAction.triggered.connect(lambda : self.savefile(self.getfilename(flag=2) ) )
		
		#Action du menu editer
		
		newAction = QtGui.QAction(QtGui.QIcon('defaults.png'),_fromUtf8('&Extraire les éléments'),self)
		newAction.setShortcut("Ctrl+N")
		newAction.setStatusTip(_fromUtf8("Créer une nouvelle feuille à partir de la sélection."))
		newAction.triggered.connect(self.newpage)
		
		copyAction= QtGui.QAction(QtGui.QIcon('copy.png'),'Copier',self)
		copyAction.setShortcut('Ctrl+C')
		copyAction.setStatusTip( _fromUtf8(u"Copier les données sélectionnées dans le presse-papier") )
		copyAction.triggered.connect(self.copycells)

		
		#Actions du menu recherche
		
		lemmaSearch = QtGui.QAction(QtGui.QIcon('searchMenu.png'),'&Recherche par lemmes',self)
		lemmaSearch.setShortcut('Ctrl+Maj+L')
		lemmaSearch.setStatusTip(_fromUtf8(u"Rechercher les différentes formes d'un mot dans le corpus"))
		lemmaSearch.triggered.connect(self.lemmasearch)
		
		#Actions du menu classifier
		classAction = QtGui.QAction(QtGui.QIcon('defaults.png'),'&Classification LDA',self)
		classAction.setShortcut('Ctrl+Maj+T')
		classAction.setStatusTip(_fromUtf8(u"Classifier automatiquement les éléments sélectionnés"))
		classAction.triggered.connect(self.classify)
		
		codAction = QtGui.QAction(QtGui.QIcon('defaults.png'),_fromUtf8('&Coder la sélection'),self)
		codAction.setShortcut('Ctrl+B')
		codAction.setStatusTip( _fromUtf8(u"Attribuer un code aux éléments sélectionnés.") )
		codAction.triggered.connect(self.codelems)
		
		
		
		#Mise en place de la barre d'outil et de la barre d'état.
		self.statusBar()
		
		#Barre d'outil
		menubar = self.menuBar()
		
		#Menu fichier
		fileMenu = menubar.addMenu('&Fichier')
		fileMenu.addAction(loadAction)
		fileMenu.addAction(saveAction)
		fileMenu.addAction(exitAction)
		
		#Menu éditer
		editMenu = menubar.addMenu(_fromUtf8("&Editer"))
		editMenu.addAction(copyAction)
		editMenu.addAction(newAction)
		
		#Menu recherche
		searchMenu = menubar.addMenu('&Recherche')
		searchMenu.addAction(lemmaSearch)
		
		#Menu classifier
		classMenu = menubar.addMenu("Classifier")
		classMenu.addAction(classAction)
		classMenu.addAction(codAction)
		
		map(self.addAction, [newAction,loadAction,saveAction,exitAction,lemmaSearch,classAction,codAction])
		
		self.tabindex=0
		self.nameindex=dict()
		self.tabs	= QtGui.QTabWidget()
		self.setCentralWidget(self.tabs)
		self.tabtable = defaultdict(dict)
		
		
		bardoutil= QToolBar()
		self.addToolBar(Qt.TopToolBarArea, bardoutil)
		
		#Paramètres du classifieur LDA
		self.params={'topics':20,
		'passes':15,
		'seuilmin':0.3,
		'maxpres':0.95,
		'minpres':2,
		'alpha':'auto',
		'eta':'auto'}
		
		
		self.ltok = Lemmtok(LEFFFPATH)
		
		
		# self.show()
	def graphicalerrors(func):
		"""
			Décorateur pour faire remonter les exceptions et les afficher dans une messagebox.
			
		"""
		@functools.wraps(func)
		def wrapper(self,*args,**kwargs):
			try:
				op=func(self,*args,**kwargs)
			except Exception as e:
				# e.__doc__
				# e.message
				logging.error(traceback.format_exc())
				errorstring=u""
				if e.__doc__:
					errorstring+=e.__doc__+"\n"
				if e.message:
					errorstring+=e.message
				QMessageBox.warning(self, _fromUtf8("Erreur"),_fromUtf8(errorstring))
				op=None
			
			return op
		
		return wrapper
	
	@graphicalerrors
	def getcurrenttab(self):
		currenttab=self.tabs.currentWidget()
		currtable=currenttab.currentWidget() if isinstance(currenttab,QTabWidget) else currenttab
		return currtable
		
	
	@graphicalerrors
	def copycells(self,*args):
		clipboard=""
		currtable=self.getcurrenttab()
		try:
			selection=currtable.selectedItems()
		except AttributeError as e:
			raise AttributeError(u"Aucune feuille ouverte !")
		
		if selection:
			minrow=min([i.row() for i in selection])
			maxrow=max([i.row() for i in selection])
			mincol=min([i.column() for i in selection])
			maxcol=max([i.column() for i in selection])
			for row in xrange(minrow,maxrow+1):
				itemlist=[]
				for col in xrange(mincol,maxcol+1):
					v = currtable[row,col] if currtable.item(row,col) in selection else ""
					itemlist.append(v)
				clipboard += "\t".join(itemlist)+"\n"
					
			# print clipboard.encode("utf-8")
			self.sys_clip.setText(clipboard)
	
	@graphicalerrors
	def getfilename(self,flag=None):
		"""
			Fonction pour obtenir le nom d'un fichier, 
			Soit à ouvrir - flag=1 
			Soit à enregistrer - flag=2 
			La fonction ouvre le dialogue par défaut du système d'exploitation.
		"""
		files_types = "XLSX (*.xlsx);;CSV (*.csv);;Tous les fichiers (*)"
		filename=""
		if flag == 1:
			filename = QFileDialog.getOpenFileName(None,
							caption=QtCore.QString("Ouvrir"),
							directory=QtCore.QString('./'),
							filter=files_types,
							selectedFilter=QtCore.QString('*.xlsx'))
		elif flag == 2:
			filename = QFileDialog.getSaveFileName(None,
							caption=QtCore.QString("Sauvegarder"),
							directory=QtCore.QString('./'),
							filter=files_types,
							selectedFilter=QtCore.QString('*.xlsx'))
		else:
			QMessageBox.critical(self,"Erreur","Le drapeau :",flag," ne correspond à aucune action.")
		return unicode(filename)
	
	@graphicalerrors
	def savefile(self,filename):
		"""
		Fonction de sauvegarde des fichiers : sauvegarde soit au format XLSX, auquel cas tout les onglets sont sauvegardés comme des feuilles
		Soit au format CSV auquel cas seul l'onglet courant est sauvegardé
		"""
		if filename:
			if filename.endswith(".xlsx"):
				currwidg=self.tabs.currentWidget()
				
				name=self.nameindex[self.tabs.currentIndex()]
				
				fichier=opx.Workbook(write_only=True,guess_types=True)
				
				if isinstance(currwidg,QTabWidget):
					for index in range(currwidg.count()):
						table=currwidg.widget(index)
						sheet=fichier.create_sheet()
						elem = unicode(currwidg.tabText(index))
						# print elem.encode("utf-8")
						sheet.title = elem
						
						# print name.encode("utf-8"),elem.encode("utf-8")
						for row in xrange(0,table.rowCount()):
							rangee=[]
							for col in xrange(0,table.columnCount()):
								v= table[row,col] if table[row,col] else None
								rangee.append(v)
							sheet.append(rangee)
							#for col in table.columnCount():
							
				elif isinstance(currwidg,QTableWidget):
					sheet=fichier.create_sheet()
					elem = name
					# print elem.encode("utf-8")
					sheet.title = elem
						
					# print name.encode("utf-8"),elem.encode("utf-8")
					for row in xrange(0,table.rowCount()):
						rangee=[]
						for col in xrange(0,table.columnCount()):
							v= table[row,col] if table[row,col] else None
							rangee.append(v)
						sheet.append(rangee)
				else:
					raise Exception("L'impossible est arrivé.")
				fichier.save(filename)
				
				message=u"Fichier enregistré :"+filename
				logging.info(message.encode("utf-8"))
				
			elif filename.endswith(".csv"):
				ok,delimiter,quotechar=self.csvaskbox()
				if ok:
					currtable=self.getcurrenttab()
					with open(filename,"w") as sortie:
						
						for row in xrange(0,currtable.rowCount()):
							for col in xrange(0,currtable.columnCount()):
								rangee=""
								element= currtable[row,col]
								
								if col > 0:
										rangee += delimiter
								if quotechar:
									rangee += quotechar + re.sub(r"("+quotechar+")","\\\1",element) + quotechar
								else:
									rangee += element
								
								sortie.write(rangee.encode("utf-8") )
								#delimiter.join([ quotechar+re.sub(r"("+quotechar+")","\\\1",currtable[row,col])+quotechar if len(quotechar) > 1 else currtable[row,col]  ]).encode("utf-8") 
							sortie.write("\n")
						
				message=u"Fichier enregistré :"+filename
				logging.info(message.encode("utf-8"))
			
			else:
				QMessageBox.warning(self, "Erreur","Format de fichier non pris en charge.")
			
		else:
			pass
	
	@graphicalerrors
	def openfile(self, filename):
		"""
			Fonction d'ouverture des fichiers.
			Ouvre les fichiers XLSX avec le module openpyxl
			Ou les fichiers CSV avec le module csv. 
			Crée un onglet pour chaque fichier ouvert, et un sub-onglet pour chaque feuille des fichiers XLSX.
		"""
		if filename:
			if filename.endswith(".xlsx"):
				subtab= QtGui.QTabWidget()
				wb = opx.load_workbook(filename, guess_types=False)
		
				for sheet in wb:
					self.tabtable[filename][sheet.title] = Table(sheet)
					subtab.addTab(self.tabtable[filename][sheet.title],sheet.title)
								
					self.tabs.addTab(subtab,os.path.basename(filename))
					self.tabs.setTabToolTip (self.tabindex, QString(filename))
					self.nameindex[self.tabindex]=filename
					self.tabindex += 1
					self.tabs.setCurrentWidget(subtab)
						# self.show()
				message=u"Ouverture du fichier "+filename
				logging.info(message.encode("utf-8"))
				
			elif filename.endswith(".csv"):
				ok,delim,qc=self.csvaskbox()
					
				if ok:
					self.tabtable[filename][os.path.basename(filename)]=Table()
					csvtable=self.tabtable[filename][os.path.basename(filename)]
					self.tabs.addTab(csvtable,os.path.basename(filename))
					self.tabs.setTabToolTip (self.tabindex, QString(filename))
					self.nameindex[self.tabindex]=filename
					self.tabindex += 1
					self.tabs.setCurrentWidget(csvtable)
					
					with open(filename) as openfile:
						z=csv.reader(openfile,delimiter=delim,quotechar=qc)
					
						
						for i,x in enumerate(z):
							if i >= csvtable.rowCount():
									rowPosition = csvtable.rowCount()
									csvtable.insertRow(rowPosition)
							for j,y in enumerate(x):
								if j >= csvtable.columnCount():
									columnPosition = csvtable.columnCount()
									csvtable.insertColumn(columnPosition)
								csvtable[i,j]=y.decode("utf-8")
							
					#subtab.addTab(self.tabtable[filename],os.path.basename(filename))
					
					message="Ouverture du fichier "+filename
					logging.info(message.encode("utf-8"))		
				
			else:
				QMessageBox.warning(self, "Erreur","Format de fichier non pris en charge.")
				message="Echec de l'ouverture du fichier "+filename
				logging.warning(message.encode("utf-8") )
	
	@graphicalerrors	
	def csvaskbox(self):
		"""
		Fonction pour demander à l'utilisateur via un QDialog les paramètres du fichier CSV à prendre en entrée. 
		Les séparateurs possibles sont la tabulation, la virgule, et le point-virgule. Les caractère de citation sont les guillemets doubles, simple.
		Ces paramètres sont transmis tels quels au module csv.
		
		"""
		
		#Dictionnaire de variable
		#Utilisé parce que les events .connect doivent être liés à des fonctions, parce qu'on ne peut pas faire d'assignation dans un lambda
		#Et parce que les fonctions internes ne peuvent pas modifier les variables de la fonction qui les enveloppe
		#Sauf en python 3 avec le mot-clef nonlocal.
		dc={'delimiter':",",
		'quotechar':"\""}
		
		box = QDialog()
		
		def chosen():
			box.accept()
		
		def die():
			box.reject()
		
		def quotecharchanged(index):
			dc['quotechar'] = ['"',"'",None][index]
		
		def delimiterchanged(index):
			dc['delimiter'] = ["\t", ",", ";"][index]
		
		
		layout = QVBoxLayout()
		lt = QHBoxLayout()
		ht = QHBoxLayout()
		tt = QHBoxLayout()
		
		delimiterchoice = QtGui.QComboBox(box)
		delimiterchoice.addItems(["Tabulation","Virgule ,","Point-Virgule ;"])
		delimiterchoice.currentIndexChanged.connect(delimiterchanged)
		# delimiterchoice.setStatusTip( _fromUtf8("Choix du caractère de séparation des valeurs dans le fichier CSV") )
		delimiterlabel = QLabel(_fromUtf8("Caractère séparateur"),box)
		
		quotecharchoice = QtGui.QComboBox(box)
		quotecharchoice.addItems(["Guillemets doubles \" ","Guillemets simples ' ","Aucun"])
		quotecharchoice.currentIndexChanged.connect(quotecharchanged)
		# quotechar
		quotecharlabel = QLabel(_fromUtf8("Caractère de citation"),box)
		
		okbutton = QPushButton("ok",box)
		okbutton.clicked.connect(chosen)
		
		cancelbutton = QPushButton("Annuler",box)
		cancelbutton.clicked.connect(die)
		
		
		lt.addWidget(delimiterlabel)
		lt.addWidget(delimiterchoice)
		lt.addWidget(quotecharlabel)
		lt.addWidget(quotecharchoice)
		
		
		ht.addWidget(okbutton)
		ht.addWidget(cancelbutton)
		
		
		
		layout.addLayout(tt)
		layout.addLayout(lt)
		layout.addLayout(ht)
		# b1.move(50,50)
		
		box.setLayout(layout)
		box.setWindowTitle(_fromUtf8(u"Paramètres du fichier CSV"))
		posx,posy=self.geometry().x(),self.geometry().y()
		box.setGeometry(posx+70,posy+80,300,50)
		box.setWindowModality(Qt.ApplicationModal)
		if box.exec_():
			return True,dc['delimiter'],dc['quotechar']
		else:
			return False, dc['delimiter'],dc['quotechar']
	
	@graphicalerrors
	def lemmasearch(self,*args):
		
		searchbox = QDialog()
		#searchbox.setWindowModality()
		searchbox.setModal(False)
		posx,posy=self.geometry().x(),self.geometry().y()
		searchbox.setGeometry(posx+90,posy+80,300,60)
		# searchbox.statusBar()
		
		def exitsearch(*args):
				searchbox.accept()
		
		# def changesearchterm():
		
		
		tokinput = QLineEdit()
		choice=QComboBox()
		c=dict()
		c['casse']=False
		c['choix']=None
		
		def handler(xs,index):
			c['choix']=xs[index]
			
		def searchlemmas(requete):
			corpus=None
			
			currtable = self.getcurrenttab()
			if currtable.selectedItems():
				corpus=currtable.selectedItems
			else:
				corpus=currtable.iteritems
			
			for item in corpus():
				if item:
					words=self.ltok.toklemize(unicode(item.text()),prune_stopwords=False,case_sensitive=c['casse'])
					
					if requete in words:
						
						currtable.setItemSelected(item,True)
					else:
						currtable.setItemSelected(item,False)
			
			
			choice.hide()
			tokinput.show()
			c['choix']=None
		
		def expl(x):
			z={"v":u"verbe" ,
				"np":u"nom propre",
				"nc":u"nom commun",
				"adj":u"adjectif",
				":GA":u":GA",
				":GN":u":GN",
				":GP":u":GP",
				":GR":u":GR",
				":NV":u":NV",
				":PV":u":PV",
				"adj":u"adjectif",
				"adjPref":u"prefixe adjectival",
				"adv":u"adverbe",
				"advneg":u"adverbe de negation",
				"advPref":u"prefixe adverbial",
				"auxAvoir":u"auxiliaire avoir",
				"auxEtre":u"auxiliaire etre",
				"caimp":u"demonstratif ça",
				"ce":u"ce",
				"cla":u"clitique accusatif",
				"clar":u"clitique accusatif reflexif",
				"cld":u"clitique direct",
				"cldr":u"clitique direct reflexif",
				"clg":u"clitique circonstant",
				"cll":u"clitique LLL",
				"cln":u"clitique nominatif",
				"clneg":u"clitique",
				"clr":u"clitique reflexif",
				"coo":u"conjonction de coordination",
				"csu":u"complementeur",
				"det":u"determinant",
				"epsilon":u"epsilon",
				"etr":u"etre",
				"GA:":u"GA:",
				"GN:":u"GN:",
				"GP:":u"GP:",
				"GR:":u"GR:",
				"ilimp":u"ilimp",
				"meta":u"meta",
				"NV:":u"NV:u",
				"parentf":u"",
				"parento":u"",
				"poncts":u"ponctuation",
				"ponctw":u"ponctuation",
				"prel":u"pronom relatif",
				"prep":u"préposition",
				"pres":u"interjection",
				"pri":u"pronom relatif/interrogatif",
				"pro":u"pronom",
				"PV:":u"PV:",
				"que":u"que",
				"que_restr":u"que",
				"sbound":u"sbound",
				"suffAdj":u"suffixe adjectival"
				}
			deb,fin=x.split(".")
			
			if fin in z:
				return deb+" ("+z[fin]+")"
			else:
				return deb
			
		def lemmafinder():
			if c['choix']:
				searchlemmas(c['choix'])
			else:
				searchterm=self.ltok.toklemize(unicode(tokinput.text()),prune_stopwords=False,case_sensitive=c['casse'])
				if searchterm:
					if len(searchterm) > 1:
						tokinput.hide()
						choice.addItems([ _fromUtf8(expl(x)) for x in searchterm])
						choice.currentIndexChanged.connect(lambda x : handler(searchterm,x) )
						choice.show()
					else:
						searchlemmas(searchterm[0])
					
				else:
					QMessageBox.warning(self,"Erreur lors de la recherche", _fromUtf8("Erreur : cette forme n'est associée à aucun lemme") )
			
		
		toklayout=QHBoxLayout()
		toktexte=QLabel("Entrer un mot :")
		toklayout.addWidget(tokinput)
		toklayout.addWidget(choice)
		
		okbutton = QPushButton("ok",searchbox)
		okbutton.clicked.connect(lemmafinder)
		
		fermer = QPushButton(_fromUtf8("Fermer"),searchbox)
		fermer.clicked.connect(exitsearch)
		
		buttonlayout = QHBoxLayout()
		buttonlayout.addWidget(okbutton)
		buttonlayout.addWidget(fermer)
		
		layout = QVBoxLayout()
		layout.addLayout(toklayout)
		layout.addLayout(buttonlayout)
		
		searchbox.setLayout(layout)
		searchbox.show()
		choice.hide()
	
	@graphicalerrors
	def classify(self,*args):
			
		
		ok,NUMTOPICS,NUMPASS,SEUILPROBA,SEUILMOT,MINIMUM,ALPHA,ETA=self.ldaaskbox()
		if ok:
			ldatable=self.newpage()
			if ldatable:
							
				corpus=dict()
				itemz=dict()
				
				
				for row in xrange(ldatable.rowCount() ):
					corpus[row]=" ".join([ldatable[row,cols] for cols in xrange(ldatable.columnCount() )])
							
				texts=self.ltok.toklemize_corpus(corpus)
				
				dictionary = corpora.Dictionary(texts)
				dictionary.filter_extremes(no_below=MINIMUM,no_above=SEUILMOT)
				dictionary.compactify()
				bow = [dictionary.doc2bow(text) for text in texts]
				corpus_tfidf =  models.TfidfModel(bow)[bow]
				lda=models.ldamodel.LdaModel(corpus=corpus_tfidf, id2word=dictionary, num_topics=NUMTOPICS,update_every=0, chunksize=4000, passes=NUMPASS, alpha=ALPHA, eta=ETA, minimum_probability=SEUILPROBA)
				
				t=dict()
				c=defaultdict(list)
				# reliquat=0
				for i in corpus:
					for topic,confid in lda[corpus_tfidf[i]]:
						t[topic]=t.get(topic,[]) + [(i,confid)]
						c[i].append( (topic,confid) )
				
				oldcount = ldatable.columnCount()
				for pos in xrange(oldcount,oldcount+NUMTOPICS):
					ldatable.insertColumn(pos)
				
				for element in corpus:
					for (i,(x,y)) in enumerate(c[element]):
						ldatable[element,oldcount+i]=str(x)
				
				codetable=Table(dimensions=(NUMTOPICS+1,3) )
				self.tabs.currentWidget().addTab(codetable,"codes")
				self.tabtable[self.nameindex[self.tabs.currentIndex()]]["codes"]=codetable
				
				for topic in sorted(t,key=lambda topic : len([x for x,y in t[topic] ]), reverse=True):
					codetable[topic,0]=str(topic)
					description=",".join([ dictionary[x].encode("utf-8") for x,y in lda.get_topic_terms(topic) ])
					codetable[topic,1]= _fromUtf8(description)
				
		
				QMessageBox.about(self, "Info",_fromUtf8("Classification automatique terminée !"))
				logging.info("Classification terminée")
	
	@graphicalerrors
	def ldaaskbox(self):
		
		alphachoix=["auto","symmetric","asymmetric"]
		etachoix=["auto",None]
		
		def changevalue(name,v):
			self.params[name]=v
		
		def changealpha(index):
			self.params['alpha']=alphachoix[index]
		
		def changeeta(index):
			self.params['eta']=etachoix[index]
		
		def chosen():
			ldabox.accept()
		
		def die():
			ldabox.reject()
		
		ldabox = QDialog()
		
		alpha = QtGui.QComboBox(ldabox)
		alpha.addItems(alphachoix)
		alpha.setCurrentIndex(alphachoix.index(self.params['alpha']))
		alpha.currentIndexChanged.connect(changealpha)
		
		
		eta = QtGui.QComboBox(ldabox)
		eta.addItems(["auto","aucun"])
		eta.setCurrentIndex(etachoix.index(self.params['eta']))
		eta.currentIndexChanged.connect(changeeta)
		
		alphalabel=QLabel( _fromUtf8(u"Alpha") )
		etalabel=QLabel( _fromUtf8(u"Eta") )
		
		ldaoptions=QHBoxLayout()
		ldaoptions.addWidget(alphalabel)
		ldaoptions.addWidget(alpha)
		ldaoptions.addWidget(etalabel)
		ldaoptions.addWidget(eta)
		
		
		
		topic=QSpinBox()
		topic.setRange(1,1000)
		topic.setValue(self.params['topics'])
		topic.valueChanged.connect(lambda x : changevalue("topics",int(x)))
		topiclabel=QLabel(_fromUtf8("Nombre de codes :"))
		topiclayer=QHBoxLayout()
		topiclayer.addWidget(topiclabel)
		topiclayer.addWidget(topic)
		
		passes=QSpinBox()
		passes.setRange(1,200)
		passes.setValue(self.params['passes'])
		passes.valueChanged.connect(lambda x : changevalue("passes",x) )
		passlabel=QLabel( _fromUtf8("Nombre de passes :") )
		passlayer=QHBoxLayout()
		passlayer.addWidget(passlabel)
		passlayer.addWidget(passes)
		
		seuil=QDoubleSpinBox()
		seuil.setRange(0,1)
		seuil.setSingleStep(0.01)
		seuil.setValue(self.params["seuilmin"])
		seuil.valueChanged.connect(lambda x : changevalue("seuilmin",x) )
		seuillabel=QLabel(_fromUtf8(u"Seuil de probabilité :"))
		seuillayer= QHBoxLayout()
		seuillayer.addWidget(seuillabel)
		seuillayer.addWidget(seuil)
		
		minmot=QSpinBox()
		minmot.setMinimum(1)
		minmot.valueChanged.connect(lambda x: changevalue("minpres",x) )
		minmot.setValue(self.params["minpres"])
		minlabel=QLabel( _fromUtf8("Hapax") )
		
		maxpres=QDoubleSpinBox()
		maxpres.setRange(0,1)
		maxpres.setSingleStep(0.01)
		maxpres.setValue(self.params["maxpres"])
		maxpres.valueChanged.connect(lambda x: changevalue('maxpres',x))
		maxlabel = QLabel(_fromUtf8("Tf max:") )
		
		motlayer=QHBoxLayout()
		motlayer.addWidget(maxlabel)
		motlayer.addWidget(maxpres)
		motlayer.addWidget(minlabel)
		motlayer.addWidget(minmot)
		
		okbutton = QPushButton("ok",ldabox)
		okbutton.clicked.connect(chosen)
		
		cancelbutton = QPushButton("Annuler",ldabox)
		cancelbutton.clicked.connect(die)
		
		buttonlayer=QHBoxLayout()
		buttonlayer.addWidget(okbutton)
		buttonlayer.addWidget(cancelbutton)
		
		layout=QVBoxLayout()
		
		layout.addLayout(topiclayer)
		layout.addLayout(passlayer)
		layout.addLayout(seuillayer)
		layout.addLayout(motlayer)
		layout.addLayout(ldaoptions)
		layout.addLayout(buttonlayer)
		
		
		ldabox.setLayout(layout)
		
		posx,posy=self.geometry().x(),self.geometry().y()
		ldabox.setGeometry(posx+90,posy+80,300,200)
		ldabox.setWindowModality(Qt.ApplicationModal)
		
		if ldabox.exec_():
			return True,self.params['topics'],self.params['passes'],self.params['seuilmin'],self.params['maxpres'],self.params['minpres'],self.params['alpha'],self.params["eta"]
		else:
			return False,0,0,0,0,0,None,None
	
	@graphicalerrors
	def codelems(self,*args):
		
		currtable=self.getcurrenttab()
		columnPosition = currtable.columnCount()
		currtable.insertColumn(columnPosition)
		selection=currtable.selectedItems()
		if selection:
			
			cbox=QDialog()
			
			c={'code':1}
			
			def codef(x):
				c['code']=x
			
			def codeit(x):
				cbox.accept()
			
			def die():
				cbox.reject()
			
			codeur=QSpinBox()
			codeur.setValue(c['code'])
			codeur.valueChanged.connect(codef)
			codlabel=QLabel( _fromUtf8(u"Code à attribuer") )
			
			okbutton =QPushButton( _fromUtf8(u"Ok"),cbox)
			okbutton.clicked.connect(codeit)
			
			layout=QVBoxLayout()
			layout.addWidget(codlabel)
			layout.addWidget(codeur)
			layout.addWidget(okbutton)
			
			cbox.setLayout(layout)
			
			if cbox.exec_():
				
				code=c['code']
				for item in selection:
					currtable[item.row(),columnPosition] = str(code)
			
		else:
			raise AttributeError(u"Sélection vide !")
	
	@graphicalerrors
	def newpage(self,*args):
		
		currtable=self.getcurrenttab()
		try:
			selection=currtable.selectedItems()
		except AttributeError as e:
			raise AttributeError(u"Aucune feuille ouverte !")
			
		if selection:
			minrow=min([i.row() for i in selection])
			maxrow=max([i.row() for i in selection])
			mincol=min([i.column() for i in selection])
			maxcol=max([i.column() for i in selection])
			
			selectname="xtal_"+str(self.new)
			self.new += 1
			
			subtab=QTabWidget()
			
			self.tabtable[selectname][selectname]=Table(dimensions=(maxrow-minrow+1,1+maxcol-mincol))
			newtable=self.tabtable[selectname][selectname]
			subtab.addTab(newtable,selectname)
			self.tabs.addTab(subtab,selectname)
			self.tabs.setTabToolTip(self.tabindex, selectname)
			self.nameindex[self.tabindex]=selectname
			self.tabindex += 1
			self.tabs.setCurrentWidget(subtab)
			for i in selection:
				newtable[i.row()-minrow, i.column() - mincol] = i.text()
			
			return newtable
		else:
			raise AttributeError(u"Sélection vide !")
				
if __name__ == '__main__':
	LEFFFPATH=os.path.dirname(os.path.realpath(sys.argv[0]))+"/lefff-3.4.mlex/lefff-3.4.mlex"
	logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)
	app = QtGui.QApplication(sys.argv)
	app.setWindowIcon(QtGui.QIcon('xtal.png'))
	ex = Main()
	sys.exit(app.exec_())
